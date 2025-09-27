import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.export_template import (
    ExportTemplate, ExportField, FieldType, FieldAlignment, GroupingType
)
from app.models.models import InvoiceData

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Motor de exportación de facturas a Excel usando templates personalizables"""
    
    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        
    def _map_alignment(self, alignment: FieldAlignment) -> str:
        """Mapea valores de FieldAlignment a valores válidos de openpyxl"""
        # Los valores del enum ya coinciden con los de openpyxl
        return alignment.value if alignment else "left"
        
    def export_invoices(self, invoices: List[InvoiceData], template: ExportTemplate) -> bytes:
        """
        Exportar facturas usando un template personalizado
        
        Args:
            invoices: Lista de facturas a exportar
            template: Template con configuración de exportación
            
        Returns:
            bytes: Archivo Excel generado
        """
        try:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            # Usar sheet_name del template o generar uno por defecto
            sheet_name = template.sheet_name or template.name[:31] or "Facturas"
            self.worksheet.title = sheet_name
            
            # Configurar estilos
            self._setup_styles()
            
            # Procesar datos según template
            processed_data = self._process_invoice_data(invoices, template)
            
            # Generar Excel
            current_row = 1
            
            # Headers
            if template.include_header:
                current_row = self._write_headers(template, current_row)
            
            # Datos
            current_row = self._write_data(processed_data, template, current_row)
            
            # Totales
            if template.include_totals:
                current_row = self._write_totals(processed_data, template, current_row)
            
            # Aplicar formato
            self._apply_formatting(template)
            
            # Guardar en memoria
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Excel generado con {len(invoices)} facturas usando template '{template.name}'")
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            raise
    
    def _setup_styles(self):
        """Configurar estilos base"""
        # Estilo para headers
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo para datos
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(vertical="center")
        
        # Bordes
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Estilo para totales
        self.total_font = Font(bold=True)
        self.total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    def _process_invoice_data(self, invoices: List[InvoiceData], template: ExportTemplate) -> List[Dict[str, Any]]:
        """
        Procesar datos de facturas según configuración del template
        
        Args:
            invoices: Facturas a procesar
            template: Template con configuración
            
        Returns:
            Datos procesados listos para escribir
        """
        processed_data = []
        
        for invoice in invoices:
            invoice_dict = invoice.model_dump()
            row_data = {}
            
            for field in sorted(template.fields, key=lambda f: f.order):
                if not field.is_visible:
                    continue
                    
                value = self._extract_field_value(invoice_dict, field)
                formatted_value = self._format_field_value(value, field)
                row_data[field.field_key] = formatted_value
                
                # Debug específico después del formateo
                if field.field_key in ['monto_exento', 'exonerado', 'total_base_gravada', 'direccion_emisor', 'telefono_emisor', 'email_emisor', 'direccion_cliente']:
                    logger.info(f"DEBUG FORMAT - Campo {field.field_key}: "
                               f"valor extraído={value} → valor formateado={formatted_value}")
            
            
            processed_data.append(row_data)
        
        return processed_data
    
    def _extract_field_value(self, invoice_dict: Dict[str, Any], field: ExportField) -> Any:
        """
        Extraer valor de un campo específico de la factura
        
        Args:
            invoice_dict: Datos de la factura
            field: Configuración del campo
            
        Returns:
            Valor extraído
        """
        try:
            # === SOLO CAMPOS REALES DE LA BD ===
            # Mapeo de compatibilidad para campos renombrados
            field_mappings = {
                "base_gravada_5": ["gravado_5", "subtotal_5"],
                "base_gravada_10": ["gravado_10", "subtotal_10"],
                # Mapeo robusto para productos: usar claves internas del producto
                "productos.nombre": ["nombre", "descripcion", "articulo"],
                "productos.descripcion": ["descripcion", "nombre", "articulo"],
            }
            
            field_key = field.field_key
            
            # Debug específico para campos problemáticos
            if field_key in ['monto_exento', 'exonerado', 'total_base_gravada']:
                logger.info(f"DEBUG EXPORT - Campo {field_key}: "
                           f"valor directo={invoice_dict.get(field_key)}, "
                           f"disponible en invoice: {field_key in invoice_dict}")
            
            # Manejo especial para campos de productos
            if field_key.startswith('productos.'):
                productos = invoice_dict.get('productos', [])
                if not productos:
                    return None
                
                # Obtener el subcampo (ej: 'productos.cantidad' -> 'cantidad')
                subcampo = field_key.split('.', 1)[1]
                
                # Usar mapeo si existe
                possible_fields = field_mappings.get(field_key, [subcampo])
                
                # Extraer valores del subcampo de todos los productos
                valores = []
                for producto in productos:
                    if isinstance(producto, dict):
                        # Intentar cada posible mapeo hasta encontrar un valor
                        valor = None
                        for posible_campo in possible_fields:
                            if posible_campo in producto:
                                valor = producto.get(posible_campo)
                                break
                        
                        # Compatibilidad adicional por si no se encontró
                        if valor is None and subcampo in ['articulo', 'descripcion', 'nombre']:
                            valor = (producto.get('nombre') or 
                                     producto.get('descripcion') or 
                                     producto.get('articulo'))
                        
                        if valor is not None:
                            valores.append(valor)
                    else:
                        # Si es un objeto Pydantic
                        valor = None
                        for posible_campo in possible_fields:
                            valor = getattr(producto, posible_campo, None)
                            if valor is not None:
                                break
                        
                        if valor is not None:
                            valores.append(valor)
                
                # Procesar como array
                if field.field_type == FieldType.ARRAY and valores:
                    return self._process_array_field(valores, field)
                
                return valores if valores else None
            
            # Para campos no anidados, usar mapeo si existe
            if field_key in field_mappings:
                for mapped_field in field_mappings[field_key]:
                    if mapped_field in invoice_dict:
                        return invoice_dict[mapped_field]
                return None
            
            # Para campos no anidados normales
            if '.' not in field_key:
                return invoice_dict.get(field_key)
            
            # Navegar por campos anidados normales
            keys = field_key.split('.')
            value = invoice_dict
            
            for key in keys:
                if isinstance(value, dict):
                    value = value.get(key)
                elif isinstance(value, list) and key.isdigit():
                    idx = int(key)
                    value = value[idx] if idx < len(value) else None
                else:
                    value = None
                    break
            
            # Manejo especial para arrays (no productos)
            if isinstance(value, list) and field.field_type == FieldType.ARRAY:
                return self._process_array_field(value, field)
            
            return value
            
        except Exception as e:
            logger.warning(f"Error extrayendo campo {field.field_key}: {e}")
            return None
    
    def _process_array_field(self, array_value: List[Any], field: ExportField) -> str:
        """
        Procesar campos de array según tipo de agrupación
        
        Args:
            array_value: Lista de valores
            field: Configuración del campo
            
        Returns:
            Valor procesado como string
        """
        if not array_value:
            return ""
        
        if field.grouping_type == GroupingType.CONCATENATE:
            # Unir todos los valores con separador, manejando diferentes tipos
            str_values = []
            for item in array_value:
                if item is not None:
                    # Para números, formatear apropiadamente
                    if isinstance(item, (int, float)):
                        if field.field_key.endswith('.cantidad'):
                            str_values.append(f"{item:.1f}")
                        elif field.field_key.endswith(('.precio_unitario', '.total')):
                            str_values.append(f"{round(item)}")  # Redondear en lugar de truncar
                        else:
                            str_values.append(str(item))
                    else:
                        str_values.append(str(item))
            return (field.separator or ", ").join(str_values)
        
        elif field.grouping_type == GroupingType.SUMMARY:
            # Para números, mostrar suma
            if field.field_type in [FieldType.NUMBER, FieldType.CURRENCY]:
                try:
                    numeric_values = [float(item) for item in array_value if item is not None]
                    total = sum(numeric_values)
                    if field.field_type == FieldType.CURRENCY:
                        return f"{round(total)}"  # Redondear totales también
                    else:
                        return f"{total:.2f}"  # Sin comas
                except:
                    pass
            # Para otros tipos, mostrar cantidad
            return f"{len(array_value)} elementos"
        
        elif field.grouping_type == GroupingType.SEPARATE_ROWS:
            # Para separate_rows, devolver el primer elemento (se manejará en escritura)
            return str(array_value[0]) if array_value else ""
        
        else:  # NONE
            return str(array_value[0]) if array_value else ""
    
    def _format_field_value(self, value: Any, field: ExportField) -> str:
        """
        Formatear valor según tipo de campo - CORREGIDO para usar redondeo apropiado
        
        Args:
            value: Valor a formatear
            field: Configuración del campo
            
        Returns:
            Valor formateado
        """
        if value is None or value == "":
            return ""
        
        try:
            if field.field_type == FieldType.CURRENCY:
                if isinstance(value, (int, float)):
                    # CORREGIDO: Usar redondeo en lugar de truncamiento
                    # round() redondea al entero más cercano (comportamiento estándar)
                    return str(round(value))
                return str(value)
            
            elif field.field_type == FieldType.NUMBER:
                if isinstance(value, (int, float)):
                    # CORREGIDO: Para cantidades, usar redondeo apropiado
                    if field.field_key in ['productos.cantidad', 'cantidad']:
                        return str(round(value))  # Redondear cantidades también
                    else:
                        return str(value)  # Otros números como están
                return str(value)
            
            elif field.field_type == FieldType.PERCENTAGE:
                if isinstance(value, (int, float)):
                    return f"{value:.1f}%"
                return str(value)
            
            elif field.field_type == FieldType.DATE:
                if isinstance(value, datetime):
                    return value.strftime("%d/%m/%Y")
                elif isinstance(value, str):
                    try:
                        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        return date_obj.strftime("%d/%m/%Y")
                    except:
                        pass
                return str(value)
            
            else:  # TEXT y otros
                return str(value)
                
        except Exception as e:
            logger.warning(f"Error formateando valor {value} para campo {field.field_key}: {e}")
            return str(value)
    
    def _write_headers(self, template: ExportTemplate, start_row: int) -> int:
        """
        Escribir fila de encabezados
        
        Args:
            template: Template con configuración
            start_row: Fila donde empezar
            
        Returns:
            Siguiente fila disponible
        """
        visible_fields = [f for f in sorted(template.fields, key=lambda f: f.order) if f.is_visible]
        
        for col_idx, field in enumerate(visible_fields, 1):
            cell = self.worksheet.cell(row=start_row, column=col_idx)
            cell.value = field.display_name
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        return start_row + 1
    
    def _write_data(self, data: List[Dict[str, Any]], template: ExportTemplate, start_row: int) -> int:
        """
        Escribir datos de facturas
        
        Args:
            data: Datos procesados
            template: Template con configuración
            start_row: Fila donde empezar
            
        Returns:
            Siguiente fila disponible
        """
        visible_fields = [f for f in sorted(template.fields, key=lambda f: f.order) if f.is_visible]
        current_row = start_row
        
        for row_data in data:
            for col_idx, field in enumerate(visible_fields, 1):
                cell = self.worksheet.cell(row=current_row, column=col_idx)
                cell.value = row_data.get(field.field_key, "")
                
                # Aplicar alineación
                alignment = Alignment(
                    horizontal=self._map_alignment(field.alignment),
                    vertical="center"
                )
                cell.alignment = alignment
                cell.border = self.thin_border
                cell.font = self.data_font
            
            current_row += 1
        
        return current_row
    
    def _write_totals(self, data: List[Dict[str, Any]], template: ExportTemplate, start_row: int) -> int:
        """
        Escribir fila de totales
        
        Args:
            data: Datos procesados
            template: Template con configuración
            start_row: Fila donde empezar
            
        Returns:
            Siguiente fila disponible
        """
        visible_fields = [f for f in sorted(template.fields, key=lambda f: f.order) if f.is_visible]
        
        # Calcular totales
        totals = {}
        for field in visible_fields:
            if field.field_type in [FieldType.CURRENCY, FieldType.NUMBER]:
                total = 0
                for row_data in data:
                    value = row_data.get(field.field_key, "")
                    try:
                        # Extraer número de string formateado
                        if isinstance(value, str):
                            clean_value = value.replace("₲", "").replace(",", "").strip()
                            total += float(clean_value) if clean_value else 0
                        elif isinstance(value, (int, float)):
                            total += value
                    except:
                        pass
                totals[field.field_key] = total
        
        # Escribir fila de totales
        for col_idx, field in enumerate(visible_fields, 1):
            cell = self.worksheet.cell(row=start_row, column=col_idx)
            
            if col_idx == 1:
                cell.value = "TOTALES"
            elif field.field_key in totals:
                if field.field_type == FieldType.CURRENCY:
                    # CORREGIDO: Usar redondeo en lugar de truncamiento para totales
                    cell.value = f"{round(totals[field.field_key])}"  # Sin comas y sin símbolo ₲
                else:
                    cell.value = f"{totals[field.field_key]:.2f}"  # Sin comas
            else:
                cell.value = ""
            
            cell.font = self.total_font
            cell.fill = self.total_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
        
        return start_row + 1
    
    def _apply_formatting(self, template: ExportTemplate):
        """
        Aplicar formato final a la hoja
        
        Args:
            template: Template con configuración
        """
        visible_fields = [f for f in sorted(template.fields, key=lambda f: f.order) if f.is_visible]
        
        # Ajustar ancho de columnas
        for col_idx, field in enumerate(visible_fields, 1):
            column_letter = get_column_letter(col_idx)
            
            if field.width:
                self.worksheet.column_dimensions[column_letter].width = field.width
            else:
                # Auto-ajustar basado en tipo de campo
                if field.field_type == FieldType.CURRENCY:
                    self.worksheet.column_dimensions[column_letter].width = 15
                elif field.field_type == FieldType.DATE:
                    self.worksheet.column_dimensions[column_letter].width = 12
                elif field.field_type == FieldType.TEXT and "ruc" in field.field_key.lower():
                    self.worksheet.column_dimensions[column_letter].width = 15
                else:
                    self.worksheet.column_dimensions[column_letter].width = 20
        
        # Freeze primera fila si hay headers
        if template.include_header:
            self.worksheet.freeze_panes = "A2"
